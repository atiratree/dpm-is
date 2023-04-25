#!/usr/bin/env python

import argparse
import os
import re

import openpyxl

validTimeSecondsRegex = re.compile("^((([0-9]|0[0-9]|1[0-9]|2[0-3]):[0-5][0-9]:[0-5][0-9])|24:00:00)$")
validTimeMinutesRegex = re.compile("^((([0-9]|0[0-9]|1[0-9]|2[0-3]):[0-5][0-9])|24:00)$")


def record_to_str(start_time, end_time, event, employee, tariff, note):
    result = ""

    def add(val):
        nonlocal result
        if val.value:
            result += " " + str(val.value)

    add(start_time)
    result += " -"
    add(end_time)
    add(event)
    add(employee)
    add(tariff)
    add(note)

    return result


def process_spreadsheet(filename, first_only, expr, check_missing_event, check_missing_employee):
    wb = openpyxl.load_workbook(filename)
    if 'Rozpis' in wb.sheetnames:
        ws = wb['Rozpis']
    else:
        ws = wb.worksheets[0]

    for row in range(1, ws.max_row):
        for day in range(0, 5):
            col_start = day * 6
            start_time = ws.cell(row=row, column=col_start + 1)
            end_time = ws.cell(row=row, column=col_start + 2)
            event = ws.cell(row=row, column=col_start + 3)
            employee = ws.cell(row=row, column=col_start + 4)
            tariff = ws.cell(row=row, column=col_start + 5)
            note = ws.cell(row=row, column=col_start + 6)

            has_start = start_time.value and (
                    validTimeSecondsRegex.match(str(start_time.value)) or validTimeMinutesRegex.match(
                str(start_time.value)))
            has_end = end_time.value and (
                    validTimeSecondsRegex.match(str(end_time.value)) or validTimeMinutesRegex.match(
                str(end_time.value)))

            # # find filled records without an event
            if check_missing_event and has_start and has_end and employee.value and tariff.value and not event.value:
                print(
                    f'Found in {filename} at row {row}; day {day}: {record_to_str(start_time, end_time, event, employee, tariff, note)}')
                if first_only:
                    return

            # # find filled records without an employee
            if check_missing_employee and has_start and has_end and event.value and tariff.value and not employee.value:
                print(
                    f'Found in {filename} at row {row}; day {day}: {record_to_str(start_time, end_time, event, employee, tariff, note)}')
                if first_only:
                    return

            # find notes with an expression
            if expr and has_start and has_end:
                m = re.search(str(expr), str(note.value), re.IGNORECASE)
                if m is not None:
                    print(f'Found in {filename} at row {row}; day {day}: {str(note.value)}')
                    if first_only:
                        return


def process(path, first_only, expr, check_missing_event, check_missing_employee):
    for root, dirs, files in os.walk(path):
        for file in files:
            filename = os.path.join(root, file)

            if filename.endswith('.xlsx'):
                process_spreadsheet(filename, first_only, expr, check_missing_event, check_missing_employee)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="path to folder with time tables", )
    parser.add_argument("--first-only", dest='first_only', action='store_true', default=False, required=False,
                        help="show only one and first result per file")
    parser.add_argument("--expr", required=False, help="expression to search for in notes column in xlsx files")
    parser.add_argument("--check-missing-event", dest='check_missing_event', action='store_true', default=False,
                        required=False,
                        help="show valid records without an event")
    parser.add_argument("--check-missing-employee", dest='check_missing_employee', action='store_true', default=False,
                        required=False,
                        help="show valid records without an employee")
    args = parser.parse_args()

    if not args.expr and not args.check_missing_event and not args.check_missing_employee:
        print("--expr or --check-missing-event or --check-missing-employee option should be specified")
        exit(1)

    process(args.path, args.first_only, args.expr, args.check_missing_event, args.check_missing_employee)
